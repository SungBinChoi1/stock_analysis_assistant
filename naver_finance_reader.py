"""
네이버증권 상승종목
- 시장경보구분 매핑
- 뉴스/공시/IR 하이퍼링크 (별도 컬럼)
- Excel 출력
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Font


class NaverStockCollectorFinal:
    """네이버증권 상승종목 수집 클래스"""

    def __init__(self):
        self.base_url = "https://stock.naver.com/api/domestic/market/stock/default"
        self.detail_url = "https://stock.naver.com/api/domestic/detail"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://stock.naver.com/market/stock/kr/stocklist/upper',
            'Accept': '*/*'
        }

        # 관리종목사유 코드-텍스트 매핑
        self.management_reason_map = {
            '6024': '자본잠식',
            '6025': '영업손실 발생',
            '6026': '매출액 미달',
            '6027': '법정관리',
            '6028': '주권 재상장',
            '6029': '지배구조 변경',
            '6051': '기타',
            '6052': '합병',
            '6053': '분할',
            '6054': '주식이전',
        }

        # 시장경보구분 코드-텍스트 매핑
        self.market_alert_map = {
            '00': '없음',
            '01': '투자주의',
            '02': '투자경고',
            '03': '투자위험'
        }

    def get_rising_stocks(self, page_size=100, start_idx=0):
        """상승종목 데이터 수집"""
        params = {
            'tradeType': 'KRX',
            'marketType': 'ALL',
            'orderType': 'up',
            'startIdx': start_idx,
            'pageSize': page_size
        }

        try:
            response = requests.get(
                self.base_url,
                headers=self.headers,
                params=params,
                timeout=10
            )
            response.raise_for_status()

            data = response.json()

            if not data:
                print("[WARNING] 데이터가 없습니다.")
                return pd.DataFrame()

            df = pd.DataFrame(data)

            # 시장 구분
            if 'sosok' in df.columns:
                df['시장구분'] = df['sosok'].map({'0': '코스피', '1': '코스닥'})

            # 상태 태그
            if 'statusTag' in df.columns:
                df['상태'] = df['statusTag'].map({
                    '0': '정상', '1': '관리종목', '2': '투자주의', '9': '거래정지'
                })

            # 등락 구분
            if 'risefall' in df.columns:
                df['등락구분'] = df['risefall'].map({
                    '1': '상한가', '2': '상승', '3': '보합', '4': '하한가', '5': '하락'
                })

            # 관리종목사유
            if 'managementReasonCode' in df.columns:
                df['관리사유'] = df['managementReasonCode'].map(
                    lambda x: self.management_reason_map.get(str(x), '') if pd.notna(x) else ''
                )

            # 시장경보
            if 'marketAlertType' in df.columns:
                df['시장경보'] = df['marketAlertType'].map(
                    lambda x: self.market_alert_map.get(str(x), '') if pd.notna(x) else ''
                )

            # 컬럼명 간소화
            columns_mapping = {
                'itemname': '종목명',
                'itemcode': '종목코드',
                'nowVal': '현재가',
                'openVal': '시가',
                'highVal': '고가',
                'lowVal': '저가',
                'changeVal': '전일대비',
                'changeRate': '등락률',
                'accQuant': '거래량',
                'accAmount': '거래대금',
                'askBuy': '매수호가',
                'askSell': '매도호가',
                'marketSum': '시가총액',
                'listedStockCnt': '상장주식수',
                'frgnRate': '외국인비율',
                'propertyTotal': '자산총계',
                'debtTotal': '부채총계',
                'sales': '매출액',
                'salesIncreasingRate': '매출증가율',
                'operatingProfit': '영업이익',
                'operatingProfitIncreasingRate': '영업이익증가율',
                'netIncome': '당기순이익',
                'eps': 'EPS',
                'per': 'PER',
                'roe': 'ROE',
                'roa': 'ROA',
                'pbr': 'PBR',
                'reserveRatio': '유보율',
                'dividend': '배당금',
                'dividendRate': '배당률',
                'high52week': '52주최고',
                'low52week': '52주최저',
                'upperLimit': '상한가',
                'lowerLimit': '하한가',
                'listedDate': '상장일',
            }

            # 필요한 컬럼만 선택
            keep_cols = list(columns_mapping.keys()) + ['시장구분', '상태', '등락구분', '관리사유', '시장경보']
            available_cols = [col for col in keep_cols if col in df.columns]
            df_selected = df[available_cols].copy()
            df_selected.rename(columns=columns_mapping, inplace=True)

            # 숫자형 변환
            numeric_cols = [
                '현재가', '시가', '고가', '저가', '전일대비', '등락률',
                '거래량', '거래대금', '매수호가', '매도호가',
                '시가총액', '상장주식수', '외국인비율',
                '자산총계', '부채총계', '매출액', '매출증가율',
                '영업이익', '영업이익증가율', '당기순이익',
                'EPS', 'PER', 'ROE', 'ROA', 'PBR', '유보율',
                '배당금', '배당률', '52주최고', '52주최저', '상한가', '하한가'
            ]

            for col in numeric_cols:
                if col in df_selected.columns:
                    df_selected[col] = pd.to_numeric(df_selected[col], errors='coerce')

            # 종목코드 A 접두사
            if '종목코드' in df_selected.columns:
                df_selected['종목코드'] = 'A' + df_selected['종목코드'].astype(str)

            df_selected['수집시간'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            print(f"[SUCCESS] {len(df_selected)}개 종목 수집 완료")
            return df_selected

        except Exception as e:
            print(f"[ERROR] 수집 실패: {e}")
            return pd.DataFrame()

    def get_stock_detail(self, stock_code):
        """종목 상세 정보"""
        clean_code = stock_code.replace('A', '')
        url = f"{self.detail_url}/{clean_code}/detail"
        params = {'codeType': 'KRX'}

        try:
            response = requests.get(url, headers=self.headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            # 종목설명 1,2,3 합치기
            descriptions = []
            for i in range(1, 4):
                desc = data.get(f'comment{i}', '')
                if desc:
                    descriptions.append(desc)

            return {
                '업종': data.get('upJongName', ''),
                '종목설명': ' '.join(descriptions)
            }
        except:
            return {'업종': '', '종목설명': ''}

    def get_recent_news(self, stock_code, months=3, max_news=3):
        """뉴스 조회 (링크 포함)"""
        clean_code = stock_code.replace('A', '')
        url = f"{self.detail_url}/news"
        params = {'itemCode': clean_code, 'page': 1, 'pageSize': 30}

        cutoff_date = datetime.now() - timedelta(days=months * 30)
        news_list = []

        try:
            response = requests.get(url, headers=self.headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            if 'clusters' in data:
                for cluster in data['clusters']:
                    for item in cluster['items']:
                        news_date_str = item.get('datetime', '')
                        if len(news_date_str) >= 8:
                            news_date = datetime.strptime(news_date_str[:8], '%Y%m%d')

                            if news_date >= cutoff_date:
                                office_id = item.get('officeId', '')
                                article_id = item.get('articleId', '')
                                link = f"https://n.news.naver.com/mnews/article/{office_id}/{article_id}"

                                news_list.append({
                                    '제목': item.get('title', ''),
                                    '일자': news_date.strftime('%Y-%m-%d'),
                                    '링크': link
                                })

                                if len(news_list) >= max_news:
                                    break

                    if len(news_list) >= max_news:
                        break
        except:
            pass

        return news_list

    def get_recent_notices(self, stock_code, months=3, max_notices=3):
        """공시 조회 (링크 포함)"""
        clean_code = stock_code.replace('A', '')
        url = f"{self.detail_url}/notice"
        params = {'itemCode': clean_code, 'startIdx': 0, 'pageSize': 30}

        cutoff_date = datetime.now() - timedelta(days=months * 30)
        notice_list = []

        try:
            response = requests.get(url, headers=self.headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            if isinstance(data, list):
                for item in data:
                    notice_date_str = item.get('datetime', '')
                    if notice_date_str:
                        try:
                            notice_date = datetime.fromisoformat(notice_date_str.replace('Z', '+00:00'))

                            if notice_date >= cutoff_date:
                                # DART 공시 링크 생성
                                biz_seq = item.get('bizdateSeq', '')
                                if '_' in biz_seq:
                                    link = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={biz_seq.split('_')[0]}"
                                else:
                                    link = ""

                                notice_list.append({
                                    '제목': item.get('title', ''),
                                    '일자': notice_date.strftime('%Y-%m-%d'),
                                    '링크': link
                                })

                                if len(notice_list) >= max_notices:
                                    break
                        except:
                            pass
        except:
            pass

        return notice_list

    def get_recent_ir(self, stock_code, months=3, max_ir=3):
        """IR 조회"""
        clean_code = stock_code.replace('A', '')
        url = f"{self.detail_url}/ir"
        params = {'itemCode': clean_code, 'startIdx': 0, 'pageSize': 30}

        cutoff_date = datetime.now() - timedelta(days=months * 30)
        ir_list = []

        try:
            response = requests.get(url, headers=self.headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            if isinstance(data, list):
                for item in data:
                    ir_date_str = item.get('writeDate', '')
                    if len(ir_date_str) >= 8:
                        try:
                            ir_date = datetime.strptime(ir_date_str[:8], '%Y%m%d')

                            if ir_date >= cutoff_date:
                                # IR 링크는 보통 없음
                                ir_list.append({
                                    '제목': item.get('title', ''),
                                    '일자': ir_date.strftime('%Y-%m-%d'),
                                    '링크': ''
                                })

                                if len(ir_list) >= max_ir:
                                    break
                        except:
                            pass
        except:
            pass

        return ir_list

    def collect_all_data(self, max_stocks=1000, include_detail=True, include_news=True):
        """전체 데이터 수집 (자동으로 모든 상승종목 수집)"""
        all_data = []
        start_idx = 0
        page_size = 100
        total_collected = 0

        print(f"\n[INFO] 상승종목 수집 시작 (최대 {max_stocks}개)")

        # 1. 기본 데이터 - 빈 결과가 나올 때까지 계속 수집
        while total_collected < max_stocks:
            print(f"[INFO] {start_idx+1}~{start_idx+page_size} 수집 중...")
            df = self.get_rising_stocks(page_size=page_size, start_idx=start_idx)

            if df.empty:
                print(f"[INFO] 더 이상 데이터가 없습니다.")
                break

            all_data.append(df)
            total_collected += len(df)

            # 받은 데이터가 page_size보다 적으면 마지막 페이지
            if len(df) < page_size:
                print(f"[INFO] 마지막 페이지 도달")
                break

            start_idx += page_size
            time.sleep(0.5)

        if not all_data:
            print("[ERROR] 수집된 데이터가 없습니다.")
            return pd.DataFrame()

        result_df = pd.concat(all_data, ignore_index=True)
        print(f"\n[SUCCESS] 총 {len(result_df)}개 상승종목 수집 완료")

        # 2. 상세정보
        if include_detail:
            print(f"\n[INFO] 상세정보(업종/설명) 수집 중...")
            details = []
            for idx, row in result_df.iterrows():
                detail = self.get_stock_detail(row['종목코드'])
                details.append(detail)

                if (idx + 1) % 10 == 0:
                    print(f"  [{idx+1}/{len(result_df)}]")
                time.sleep(0.2)

            detail_df = pd.DataFrame(details)
            result_df = pd.concat([result_df, detail_df], axis=1)
            print(f"[SUCCESS] 상세정보 완료")

        # 3. 뉴스/공시/IR (각각 별도 컬럼으로)
        if include_news:
            print(f"\n[INFO] 뉴스/공시/IR 수집 중... (시간이 걸립니다)")

            for idx, row in result_df.iterrows():
                stock_code = row['종목코드']

                # 뉴스
                news = self.get_recent_news(stock_code)
                for i in range(3):
                    if i < len(news):
                        result_df.at[idx, f'뉴스{i+1}'] = news[i]['제목'][:50]
                        result_df.at[idx, f'뉴스{i+1}_일자'] = news[i]['일자']
                        result_df.at[idx, f'뉴스{i+1}_링크'] = news[i]['링크']
                    else:
                        result_df.at[idx, f'뉴스{i+1}'] = ''
                        result_df.at[idx, f'뉴스{i+1}_일자'] = ''
                        result_df.at[idx, f'뉴스{i+1}_링크'] = ''

                # 공시
                notices = self.get_recent_notices(stock_code)
                for i in range(3):
                    if i < len(notices):
                        result_df.at[idx, f'공시{i+1}'] = notices[i]['제목'][:50]
                        result_df.at[idx, f'공시{i+1}_일자'] = notices[i]['일자']
                        result_df.at[idx, f'공시{i+1}_링크'] = notices[i]['링크']
                    else:
                        result_df.at[idx, f'공시{i+1}'] = ''
                        result_df.at[idx, f'공시{i+1}_일자'] = ''
                        result_df.at[idx, f'공시{i+1}_링크'] = ''

                # IR
                ir_list = self.get_recent_ir(stock_code)
                for i in range(3):
                    if i < len(ir_list):
                        result_df.at[idx, f'IR{i+1}'] = ir_list[i]['제목'][:50]
                        result_df.at[idx, f'IR{i+1}_일자'] = ir_list[i]['일자']
                    else:
                        result_df.at[idx, f'IR{i+1}'] = ''
                        result_df.at[idx, f'IR{i+1}_일자'] = ''

                if (idx + 1) % 5 == 0:
                    print(f"  [{idx+1}/{len(result_df)}]")
                time.sleep(0.3)

            print(f"[SUCCESS] 뉴스/공시/IR 완료")

        return result_df

    def save_to_excel_with_hyperlinks(self, df, filename=None, output_dir='data'):
        """하이퍼링크가 적용된 Excel 저장"""
        if df.empty:
            print("[WARNING] 저장할 데이터가 없습니다.")
            return

        os.makedirs(output_dir, exist_ok=True)

        if filename is None:
            today = datetime.now().strftime('%Y%m%d')
            filename = f'rising_stocks_{today}.xlsx'

        filepath = os.path.join(output_dir, filename)

        # Excel로 저장
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='상승종목')

        # 하이퍼링크 적용
        wb = load_workbook(filepath)
        ws = wb['상승종목']

        # 열 인덱스 찾기
        header = [cell.value for cell in ws[1]]

        # 뉴스/공시 하이퍼링크 적용
        for row_idx in range(2, ws.max_row + 1):
            for col_name in ['뉴스1', '뉴스2', '뉴스3', '공시1', '공시2', '공시3']:
                if col_name in header:
                    col_idx = header.index(col_name) + 1
                    link_col_idx = header.index(f'{col_name}_링크') + 1

                    title_cell = ws.cell(row=row_idx, column=col_idx)
                    link_cell = ws.cell(row=row_idx, column=link_col_idx)

                    if link_cell.value:
                        title_cell.hyperlink = link_cell.value
                        title_cell.font = Font(color="0563C1", underline="single")

        # 링크 컬럼 숨기기
        for col_name in header:
            if '_링크' in col_name:
                col_idx = header.index(col_name) + 1
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].hidden = True

        # 열 너비 조정
        for column in ws.columns:
            if not ws.column_dimensions[column[0].column_letter].hidden:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(filepath)

        print(f"\n[SAVED] {filepath}")
        print(f"        종목: {len(df)}개")
        print(f"        컬럼: {len(df.columns)}개")
        print(f"        크기: {os.path.getsize(filepath) / 1024:.2f} KB")

        return filepath


def main():
    """메인 실행"""
    print("네이버증권 종목분석\n")

    collector = NaverStockCollectorFinal()

    print("="*80)
    print("수집 옵션")
    print("="*80)
    print("1. 기본 (~10초) - 가격/재무 정보만")
    print("2. 기본+상세 (~5분) - 업종/종목설명 포함")
    print("3. 전체 (~15분) - 뉴스/공시/IR 포함")
    print("\n※ 모든 상승종목을 자동으로 수집합니다 (100개 이상 가능)")

    choice = input("\n선택 (1/2/3, 기본=2): ").strip() or "2"

    include_detail = choice in ["2", "3"]
    include_news = choice == "3"

    # 수집 (자동으로 모든 상승종목 수집)
    df = collector.collect_all_data(
        max_stocks=1000,  # 최대 1000개까지
        include_detail=include_detail,
        include_news=include_news
    )

    if not df.empty:
        # 저장
        filepath = collector.save_to_excel_with_hyperlinks(df)

        print(f"\n 완료 {filepath}")
        print(f"   총 {len(df)}개 종목 수집")
        print("\n[참고] 뉴스/공시 제목을 클릭하면 링크로 이동합니다.")
        print("[참고] 대시보드 실행: streamlit run stock_dashboard.py")
    else:
        print("\n수집 실패")


if __name__ == "__main__":
    main()