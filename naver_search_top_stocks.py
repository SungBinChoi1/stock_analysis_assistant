"""
네이버증권 구버전 - 검색상위종목 (전일비 기준 상승종목)
URL: https://finance.naver.com/sise/lastsearch2.naver
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class NaverSearchTopStocks:
    """네이버증권 구버전 검색상위종목 크롤러"""
    
    def __init__(self):
        self.base_url = "https://finance.naver.com/sise/lastsearch2.naver"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Referer': 'https://finance.naver.com/',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7'
        }
    
    def clean_text(self, text):
        """텍스트 정리"""
        if text:
            return text.strip().replace('\n', '').replace('\t', '').replace(',', '')
        return ''
    
    def parse_number(self, text):
        """숫자 파싱"""
        try:
            cleaned = self.clean_text(text)
            if cleaned == '' or cleaned == 'N/A':
                return None
            # + / - / % 기호 제거하고 숫자만 추출
            cleaned = cleaned.replace('+', '').replace('-', '').replace('%', '')
            return float(cleaned)
        except:
            return None
    
    def get_search_top_stocks(self, sort='changeRate', ascending=False):
        """
        검색상위종목 크롤링
        
        Parameters:
        -----------
        sort : str
            정렬 기준 ('changeRate', 'stockName', 'nowVal', 'changeVal', 'quant', 'amount')
            - changeRate: 전일비 (기본값)
            - stockName: 종목명
            - nowVal: 현재가
            - changeVal: 전일대비
            - quant: 거래량
            - amount: 거래대금
        ascending : bool
            오름차순 여부 (False=내림차순, True=오름차순)
        """
        
        # 정렬 파라미터 매핑
        sort_map = {
            'changeRate': 'changeRate',  # 전일비
            'stockName': 'stockName',     # 종목명
            'nowVal': 'nowVal',           # 현재가
            'changeVal': 'changeVal',     # 전일대비
            'quant': 'quant',             # 거래량
            'amount': 'amount'            # 거래대금
        }
        
        params = {
            'sosok': '0',  # 0: 전체, 1: 코스피, 2: 코스닥
            'type': '1'    # 1: 전일비순
        }
        
        try:
            print(f"\n[INFO] 검색상위종목 수집 시작...")
            print(f"[INFO] URL: {self.base_url}")
            
            response = requests.get(
                self.base_url,
                headers=self.headers,
                params=params,
                timeout=15
            )
            response.raise_for_status()
            response.encoding = 'euc-kr'  # 네이버 구버전 인코딩
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # 디버깅: 페이지 구조 확인
            print(f"[DEBUG] 페이지 제목: {soup.title.string if soup.title else 'None'}")
            
            # 데이터 테이블 찾기 (여러 방법 시도)
            table = soup.find('table', {'class': 'type_5'})
            
            if not table:
                # 다른 클래스명 시도
                table = soup.find('table', {'summary': '검색 상위 종목'})
            
            if not table:
                # 모든 table 태그 확인
                all_tables = soup.find_all('table')
                print(f"[DEBUG] 페이지 내 테이블 개수: {len(all_tables)}")
                
                if all_tables:
                    for idx, t in enumerate(all_tables):
                        print(f"[DEBUG] 테이블 {idx}: class={t.get('class')}, summary={t.get('summary')}")
                    
                    # 가장 큰 테이블을 선택
                    table = max(all_tables, key=lambda t: len(str(t)))
                    print(f"[INFO] 가장 큰 테이블 선택")
            
            if not table:
                print("[ERROR] 테이블을 찾을 수 없습니다.")
                print("[INFO] HTML 샘플:")
                print(response.text[:1000])
                return pd.DataFrame()
            
            # tbody 확인
            tbody = table.find('tbody')
            if not tbody:
                print("[WARNING] tbody 없음, table에서 직접 tr 추출")
                rows = table.find_all('tr')
            else:
                rows = tbody.find_all('tr')
            
            # 데이터 수집
            stocks_data = []
            
            for row in rows:
                cols = row.find_all('td')
                
                # td가 없거나, 헤더 행, 빈 공백 행은 스킵
                if not cols or row.get('class') == ['type1']:
                    continue
                
                # colspan이 있는 빈 행 스킵
                if len(cols) == 1 and cols[0].get('colspan'):
                    continue
                
                if len(cols) < 11:
                    continue
                
                try:
                    # 순위 (0번째)
                    rank = self.clean_text(cols[0].text) if len(cols) > 0 else ''
                    
                    # 종목명과 종목코드 (1번째)
                    stock_link = cols[1].find('a') if len(cols) > 1 else None
                    stock_name = self.clean_text(stock_link.text) if stock_link else ''
                    stock_code = ''
                    if stock_link and stock_link.get('href'):
                        href = stock_link.get('href')
                        if 'code=' in href:
                            stock_code = 'A' + href.split('code=')[1].split('&')[0]
                    
                    # 검색비율 (2번째)
                    search_ratio = self.parse_number(cols[2].text) if len(cols) > 2 else None
                    
                    # 현재가 (3번째)
                    current_price = self.parse_number(cols[3].text) if len(cols) > 3 else None
                    
                    # 전일대비 (4번째) - span 클래스로 상승/하락 구분
                    change_val = None
                    change_type = '보합'
                    if len(cols) > 4:
                        change_col = cols[4]
                        change_val = self.parse_number(change_col.text)
                        
                        # span 태그 찾기 (class="blind"는 제외)
                        spans = change_col.find_all('span')
                        for span in spans:
                            span_class = span.get('class', [])
                            # blind 클래스는 스킵
                            if 'blind' in span_class:
                                continue
                            
                            # 색상 클래스로 상승/하락 구분
                            # red01, red02 → 상승
                            # nv01 → 하락 (navy)
                            span_class_str = ' '.join(span_class)
                            if 'red' in span_class_str:
                                change_type = '상승'
                                break
                            elif 'nv' in span_class_str or 'blue' in span_class_str:
                                change_type = '하락'
                                break
                    
                    # 등락률 (5번째)
                    change_rate_text = cols[5].text if len(cols) > 5 else ''
                    change_rate = self.parse_number(change_rate_text)
                    
                    # 디버깅: 첫 번째 종목만 출력
                    if not stocks_data:
                        print(f"\n[DEBUG] 첫 번째 종목 파싱:")
                        print(f"  td[4] 텍스트: '{cols[4].text if len(cols) > 4 else 'N/A'}'")
                        print(f"  td[4] span: {cols[4].find('span') if len(cols) > 4 else None}")
                        print(f"  td[5] 텍스트: '{change_rate_text}'")
                        print(f"  td[5] 파싱결과: {change_rate}")
                        if len(cols) > 4:
                            span = cols[4].find('span')
                            if span:
                                print(f"  span class: {span.get('class')}")

                    
                    # 거래량 (6번째)
                    volume = self.parse_number(cols[6].text) if len(cols) > 6 else None
                    
                    # 시가 (7번째)
                    open_price = self.parse_number(cols[7].text) if len(cols) > 7 else None
                    
                    # 고가 (8번째)
                    high_price = self.parse_number(cols[8].text) if len(cols) > 8 else None
                    
                    # 저가 (9번째)
                    low_price = self.parse_number(cols[9].text) if len(cols) > 9 else None
                    
                    # PER (10번째)
                    per = self.parse_number(cols[10].text) if len(cols) > 10 else None
                    
                    # ROE (11번째)
                    roe = self.parse_number(cols[11].text) if len(cols) > 11 else None
                    
                    stock_info = {
                        '순위': rank,
                        '종목명': stock_name,
                        '종목코드': stock_code,
                        '검색비율': search_ratio,
                        '현재가': current_price,
                        '전일대비': change_val,
                        '등락률': change_rate,
                        '등락구분': change_type,
                        '거래량': volume,
                        '시가': open_price,
                        '고가': high_price,
                        '저가': low_price,
                        'PER': per,
                        'ROE': roe
                    }
                    
                    stocks_data.append(stock_info)
                    
                except Exception as e:
                    print(f"[WARNING] 행 파싱 오류: {e}")
                    continue
            
            if not stocks_data:
                print("[ERROR] 수집된 데이터가 없습니다.")
                return pd.DataFrame()
            
            # DataFrame 생성
            df = pd.DataFrame(stocks_data)
            
            # 수집시간 추가
            df['수집시간'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 디버깅: 처음 3개 종목 출력
            print("\n[DEBUG] 수집된 데이터 샘플:")
            for idx in range(min(3, len(df))):
                print(f"\n종목 #{idx+1}:")
                print(f"  종목명: {df.iloc[idx]['종목명']}")
                print(f"  등락구분: '{df.iloc[idx]['등락구분']}'")
                print(f"  등락률: {df.iloc[idx]['등락률']}")
                print(f"  현재가: {df.iloc[idx]['현재가']}")
            
            # 정렬
            if sort in df.columns:
                df = df.sort_values(by=sort, ascending=ascending, na_position='last')
                df = df.reset_index(drop=True)
            
            print(f"\n[SUCCESS] {len(df)}개 종목 수집 완료")
            
            return df
            
        except Exception as e:
            print(f"[ERROR] 크롤링 실패: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def get_rising_stocks_only(self, min_change_rate=0.0):
        """
        상승종목만 필터링
        
        Parameters:
        -----------
        min_change_rate : float
            최소 등락률 (기본값: 0.0, 즉 상승종목만)
        """
        df = self.get_search_top_stocks(sort='changeRate', ascending=False)
        
        if df.empty:
            return df
        
        # 상승종목만 필터링
        rising_df = df[
            (df['등락구분'] == '상승') & 
            (df['등락률'] >= min_change_rate)
        ].copy()
        
        rising_df = rising_df.reset_index(drop=True)
        
        print(f"\n[FILTER] 상승종목: {len(rising_df)}개 (등락률 >= {min_change_rate}%)")
        
        return rising_df
    
    def save_to_excel(self, df, filename=None, output_dir='data'):
        """Excel 저장 with 스타일링"""
        if df.empty:
            print("[WARNING] 저장할 데이터가 없습니다.")
            return
        
        os.makedirs(output_dir, exist_ok=True)
        
        if filename is None:
            today = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'search_top_stocks_{today}.xlsx'
        
        filepath = os.path.join(output_dir, filename)
        
        # Excel 저장
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='검색상위종목')
        
        # 스타일 적용
        wb = load_workbook(filepath)
        ws = wb['검색상위종목']
        
        # 헤더 스타일
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 행 스타일
        for row_idx in range(2, ws.max_row + 1):
            # 등락구분에 따른 색상 적용
            header = [cell.value for cell in ws[1]]
            
            if '등락구분' in header:
                change_type_col = header.index('등락구분') + 1
                change_type = ws.cell(row=row_idx, column=change_type_col).value
                
                if change_type == '상승':
                    fill_color = 'FFE6E6'  # 연한 빨강
                    font_color = 'FF0000'  # 빨강
                elif change_type == '하락':
                    fill_color = 'E6F0FF'  # 연한 파랑
                    font_color = '0000FF'  # 파랑
                else:
                    fill_color = 'F0F0F0'  # 회색
                    font_color = '000000'  # 검정
                
                # 전일대비, 등락률 컬럼에 색상 적용
                for col_name in ['전일대비', '등락률', '등락구분']:
                    if col_name in header:
                        col_idx = header.index(col_name) + 1
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                        cell.font = Font(color=font_color, bold=True)
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 틀 고정 (헤더)
        ws.freeze_panes = 'A2'
        
        wb.save(filepath)
        
        print(f"\n[SAVED] {filepath}")
        print(f"        종목: {len(df)}개")
        print(f"        컬럼: {len(df.columns)}개")
        print(f"        크기: {os.path.getsize(filepath) / 1024:.2f} KB")
        
        return filepath
    
    def display_summary(self, df):
        """데이터 요약 출력"""
        if df.empty:
            print("[INFO] 데이터가 없습니다.")
            return
        
        print("\n" + "="*80)
        print("수집 결과 요약")
        print("="*80)
        
        print(f"\n총 종목 수: {len(df)}개")
        
        # 등락구분별 집계
        if '등락구분' in df.columns:
            print("\n[등락구분별]")
            change_counts = df['등락구분'].value_counts()
            for change_type, count in change_counts.items():
                print(f"  {change_type}: {count}개")
        
        # 등락률 통계
        if '등락률' in df.columns:
            print("\n[등락률 통계]")
            print(f"  평균: {df['등락률'].mean():.2f}%")
            print(f"  최대: {df['등락률'].max():.2f}%")
            print(f"  최소: {df['등락률'].min():.2f}%")
        
        # 상위 5개 종목
        print("\n[등락률 상위 5개]")
        display_cols = ['순위', '종목명', '검색비율', '현재가', '전일대비', '등락률']
        available_cols = [col for col in display_cols if col in df.columns]
        top5 = df.nlargest(5, '등락률')[available_cols]
        print(top5.to_string(index=False))
        
        print("\n" + "="*80)


def main():
    """메인 실행"""
    print("="*80)
    print("네이버증권 구버전 - 검색상위종목 크롤러")
    print("="*80)
    print("\n[URL] https://finance.naver.com/sise/lastsearch2.naver")
    
    collector = NaverSearchTopStocks()
    
    print("\n수집 옵션:")
    print("1. 전체 종목 (상승/하락/보합 모두)")
    print("2. 상승종목만 (등락률 > 0%)")
    print("3. 상승종목만 (등락률 >= 5%)")
    
    choice = input("\n선택 (1/2/3, 기본=2): ").strip() or "2"
    
    if choice == "1":
        df = collector.get_search_top_stocks(sort='changeRate', ascending=False)
    elif choice == "2":
        df = collector.get_rising_stocks_only(min_change_rate=0.0)
    elif choice == "3":
        df = collector.get_rising_stocks_only(min_change_rate=5.0)
    else:
        print("[ERROR] 잘못된 선택입니다.")
        return
    
    if not df.empty:
        # 요약 출력
        collector.display_summary(df)
        
        # Excel 저장
        filepath = collector.save_to_excel(df)
        
        print(f"\n✅ 완료: {filepath}")
        print(f"   총 {len(df)}개 종목 수집")
    else:
        print("\n❌ 수집 실패")


if __name__ == "__main__":
    main()
